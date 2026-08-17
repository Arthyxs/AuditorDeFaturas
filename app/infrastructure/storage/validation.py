"""Content-aware validation for untrusted uploaded documents."""

import csv
import warnings
from dataclasses import dataclass
from pathlib import Path, PurePath
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

from defusedxml.common import DefusedXmlException  # type: ignore[import-untyped]
from defusedxml.ElementTree import iterparse  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from PIL import Image, UnidentifiedImageError
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from xlrd import open_workbook  # type: ignore[import-untyped]
from xlrd.biffh import XLRDError  # type: ignore[import-untyped]
from xlrd.compdoc import CompDocError  # type: ignore[import-untyped]

_MAX_ARCHIVE_ENTRIES = 10_000
_MAX_ARCHIVE_EXPANDED_BYTES = 256 * 1024 * 1024
_MAX_ARCHIVE_COMPRESSION_RATIO = 100
_MAX_PDF_PAGES = 10_000
_MAX_IMAGE_FRAMES = 1_000
_MAX_IMAGE_PIXELS = 40_000_000


class UploadValidationError(ValueError):
    """An upload violates the approved filename, type or size policy."""


@dataclass(frozen=True)
class UploadType:
    """Approved extension, declared MIME and canonical stored MIME."""

    declared_mimes: frozenset[str]
    canonical_mime: str


UPLOAD_TYPES: dict[str, UploadType] = {
    ".pdf": UploadType(frozenset({"application/pdf"}), "application/pdf"),
    ".xlsx": UploadType(
        frozenset({"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
    ".xls": UploadType(frozenset({"application/vnd.ms-excel"}), "application/vnd.ms-excel"),
    ".csv": UploadType(frozenset({"text/csv", "application/csv"}), "text/csv"),
    ".png": UploadType(frozenset({"image/png"}), "image/png"),
    ".jpg": UploadType(frozenset({"image/jpeg"}), "image/jpeg"),
    ".jpeg": UploadType(frozenset({"image/jpeg"}), "image/jpeg"),
    ".tif": UploadType(frozenset({"image/tiff"}), "image/tiff"),
    ".tiff": UploadType(frozenset({"image/tiff"}), "image/tiff"),
}


@dataclass(frozen=True)
class ValidatedUpload:
    """Normalized metadata after filename, MIME and content validation."""

    original_filename: str
    extension: str
    mime_type: str


class UploadValidator:
    """Validate untrusted uploads without executing or extracting their content."""

    def __init__(self, max_size_bytes: int) -> None:
        if max_size_bytes < 1:
            raise ValueError("max_size_bytes must be positive")
        self.max_size_bytes = max_size_bytes

    def validate(
        self, path: Path, *, original_filename: str, declared_mime: str, size: int
    ) -> ValidatedUpload:
        """Validate metadata and inspect the staged bytes against their extension."""
        filename = validate_original_filename(original_filename)
        if size < 1:
            raise UploadValidationError("empty files are not accepted")
        if size > self.max_size_bytes:
            raise UploadValidationError("file exceeds the configured size limit")
        if path.stat().st_size != size:
            raise UploadValidationError("staged file size changed during validation")

        extension = Path(filename).suffix.casefold()
        upload_type = UPLOAD_TYPES.get(extension)
        if upload_type is None:
            raise UploadValidationError("file extension is not supported")
        normalized_mime = declared_mime.partition(";")[0].strip().casefold()
        if normalized_mime not in upload_type.declared_mimes:
            raise UploadValidationError("declared MIME does not match the file extension")

        _validate_content(path, extension, size)
        return ValidatedUpload(
            original_filename=filename,
            extension=extension,
            mime_type=upload_type.canonical_mime,
        )


def validate_original_filename(filename: str) -> str:
    """Reject traversal, absolute paths, control characters and ambiguous names."""
    if not filename or len(filename) > 255:
        raise UploadValidationError("filename must contain 1-255 characters")
    if filename in {".", ".."} or filename != PurePath(filename).name:
        raise UploadValidationError("filename must not contain a path")
    if "/" in filename or "\\" in filename or ":" in filename:
        raise UploadValidationError("filename must not contain path separators")
    if any(ord(character) < 32 or ord(character) == 127 for character in filename):
        raise UploadValidationError("filename contains control characters")
    return filename


def _validate_content(path: Path, extension: str, size: int) -> None:
    if extension == ".pdf":
        _validate_pdf(path)
    elif extension in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
        _validate_image(path, extension)
    elif extension == ".xls":
        _validate_xls(path)
    elif extension == ".xlsx":
        _validate_xlsx(path)
    elif extension == ".csv":
        _validate_csv(path)


def _validate_pdf(path: Path) -> None:
    with path.open("rb") as uploaded:
        prefix = uploaded.read(8)
        uploaded.seek(max(0, path.stat().st_size - 4096))
        suffix = uploaded.read()
    eof_marker = suffix.rfind(b"%%EOF")
    if (
        not prefix.startswith(b"%PDF-")
        or eof_marker < 0
        or suffix[eof_marker + len(b"%%EOF") :].strip(b"\x00\t\n\r\f ")
    ):
        raise UploadValidationError("content is not a complete standalone PDF")

    try:
        reader = PdfReader(path, strict=True)
        if reader.is_encrypted:
            raise UploadValidationError("encrypted PDFs are not supported")
        page_count = len(reader.pages)
        if page_count < 1 or page_count > _MAX_PDF_PAGES:
            raise UploadValidationError("PDF page count is outside the safe range")
        for page in reader.pages:
            _ = page.mediabox
    except UploadValidationError:
        raise
    except (PdfReadError, OSError, TypeError, ValueError) as exc:
        raise UploadValidationError("content is not a parseable PDF") from exc


def _validate_image(path: Path, extension: str) -> None:
    expected_format = {
        ".png": "PNG",
        ".jpg": "JPEG",
        ".jpeg": "JPEG",
        ".tif": "TIFF",
        ".tiff": "TIFF",
    }[extension]
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(path) as image:
                if image.format != expected_format:
                    raise UploadValidationError("image content does not match its extension")
                image.verify()
            with Image.open(path) as image:
                frame_count = getattr(image, "n_frames", 1)
                if frame_count < 1 or frame_count > _MAX_IMAGE_FRAMES:
                    raise UploadValidationError("image frame count is outside the safe range")
                total_pixels = 0
                for frame_number in range(frame_count):
                    image.seek(frame_number)
                    total_pixels += image.width * image.height
                    if total_pixels > _MAX_IMAGE_PIXELS:
                        raise UploadValidationError("image dimensions exceed the safe pixel limit")
                    image.load()
    except UploadValidationError:
        raise
    except (
        EOFError,
        Image.DecompressionBombError,
        Image.DecompressionBombWarning,
        OSError,
        UnidentifiedImageError,
        ValueError,
    ) as exc:
        raise UploadValidationError("content is not a parseable image") from exc


def _validate_xls(path: Path) -> None:
    try:
        workbook = open_workbook(path, on_demand=True)
        try:
            if workbook.nsheets < 1:
                raise UploadValidationError("legacy spreadsheet contains no worksheets")
            for index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(index)
                if sheet.nrows and sheet.ncols:
                    _ = sheet.cell_value(sheet.nrows - 1, sheet.ncols - 1)
                workbook.unload_sheet(index)
        finally:
            workbook.release_resources()
    except UploadValidationError:
        raise
    except (CompDocError, EOFError, IndexError, OSError, TypeError, ValueError, XLRDError) as exc:
        raise UploadValidationError("content is not a parseable XLS spreadsheet") from exc


def _validate_xlsx(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            names = [entry.filename for entry in entries]
            if len(entries) > _MAX_ARCHIVE_ENTRIES:
                raise UploadValidationError("spreadsheet contains too many archive entries")
            if len(names) != len(set(names)):
                raise UploadValidationError("spreadsheet contains duplicate archive entries")
            if sum(entry.file_size for entry in entries) > _MAX_ARCHIVE_EXPANDED_BYTES:
                raise UploadValidationError("spreadsheet expands beyond the safe limit")
            if any(
                entry.compress_size > 0
                and entry.file_size > entry.compress_size * _MAX_ARCHIVE_COMPRESSION_RATIO
                for entry in entries
            ):
                raise UploadValidationError("spreadsheet archive compression ratio is unsafe")
            if any(
                entry.flag_bits & 1
                or entry.filename.startswith(("/", "\\"))
                or ".." in PurePath(entry.filename.replace("\\", "/")).parts
                for entry in entries
            ):
                raise UploadValidationError("spreadsheet archive paths or encryption are unsafe")
            required_parts = {
                "[Content_Types].xml",
                "_rels/.rels",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
            }
            if not required_parts.issubset(names):
                raise UploadValidationError("ZIP content is not an XLSX spreadsheet")
            if archive.testzip() is not None:
                raise UploadValidationError("spreadsheet archive is corrupt")
            for entry in entries:
                if entry.filename.endswith((".xml", ".rels")):
                    with archive.open(entry) as xml_stream:
                        for _, element in iterparse(xml_stream, events=("end",)):
                            element.clear()

        with path.open("rb") as spreadsheet:
            workbook = load_workbook(spreadsheet, read_only=True, data_only=False, keep_links=False)
            try:
                if not workbook.sheetnames:
                    raise UploadValidationError("spreadsheet contains no worksheets")
            finally:
                workbook.close()
    except UploadValidationError:
        raise
    except (
        BadZipFile,
        DefusedXmlException,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        TypeError,
        ValueError,
    ) as exc:
        raise UploadValidationError("content is not an XLSX spreadsheet") from exc


def _validate_csv(path: Path) -> None:
    with path.open("rb") as binary_file:
        for chunk in iter(lambda: binary_file.read(1024 * 1024), b""):
            if b"\x00" in chunk:
                raise UploadValidationError("CSV contains binary null bytes")
            text = chunk.decode("latin-1")
            if any(ord(character) < 32 and character not in "\t\r\n" for character in text):
                raise UploadValidationError("CSV contains binary control characters")

    try:
        with path.open("r", encoding="utf-8-sig", errors="strict", newline="") as utf8_file:
            reader = csv.reader(utf8_file)
            if next(reader, None) is None:
                raise UploadValidationError("CSV contains no records")
            for _ in reader:
                pass
    except UnicodeDecodeError:
        try:
            with path.open("r", encoding="cp1252", errors="strict", newline="") as cp1252_file:
                reader = csv.reader(cp1252_file)
                if next(reader, None) is None:
                    raise UploadValidationError("CSV contains no records")
                for _ in reader:
                    pass
        except (UnicodeDecodeError, csv.Error) as exc:
            raise UploadValidationError("content is not a parseable CSV") from exc
    except csv.Error as exc:
        raise UploadValidationError("content is not a parseable CSV") from exc
