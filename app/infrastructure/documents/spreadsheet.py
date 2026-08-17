"""Layout-neutral spreadsheet tools for XLSX, XLS and delimited text."""

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Protocol

import openpyxl  # type: ignore[import-untyped]
import xlrd  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore[import-untyped]

from app.infrastructure.documents.common import verified_bytes
from app.infrastructure.documents.models import (
    CellSearchMatch,
    CellSearchResult,
    CellValue,
    DocumentReference,
    DocumentToolError,
    EvidenceCoordinate,
    SheetInfo,
    SheetList,
    SheetRangeResult,
)
from app.ports.storage import StorageProvider


class _Sheet(Protocol):
    @property
    def name(self) -> str: ...

    @property
    def max_row(self) -> int: ...

    @property
    def max_column(self) -> int: ...

    def value(self, row: int, column: int) -> object: ...

    def formula(self, row: int, column: int) -> str | None: ...


@dataclass
class _MatrixSheet:
    name: str
    rows: list[list[object]]

    @property
    def max_row(self) -> int:
        return len(self.rows)

    @property
    def max_column(self) -> int:
        return max((len(row) for row in self.rows), default=0)

    def value(self, row: int, column: int) -> object:
        if row > len(self.rows) or column > len(self.rows[row - 1]):
            return None
        return self.rows[row - 1][column - 1]

    def formula(self, row: int, column: int) -> str | None:
        value = self.value(row, column)
        return value if isinstance(value, str) and value.startswith("=") else None


@dataclass
class _OpenpyxlSheet:
    worksheet: object

    @property
    def name(self) -> str:
        return str(self.worksheet.title)  # type: ignore[attr-defined]

    @property
    def max_row(self) -> int:
        return int(self.worksheet.max_row)  # type: ignore[attr-defined]

    @property
    def max_column(self) -> int:
        return int(self.worksheet.max_column)  # type: ignore[attr-defined]

    def value(self, row: int, column: int) -> object:
        return self.worksheet.cell(row=row, column=column).value  # type: ignore[attr-defined]

    def formula(self, row: int, column: int) -> str | None:
        value = self.value(row, column)
        return value if isinstance(value, str) and value.startswith("=") else None


class SpreadsheetTools:
    def __init__(
        self,
        storage: StorageProvider,
        *,
        max_file_bytes: int = 50 * 1024 * 1024,
        max_cells_per_call: int = 10_000,
        max_search_cells: int = 100_000,
    ) -> None:
        self._storage = storage
        self._max_file_bytes = max_file_bytes
        self._max_cells = max_cells_per_call
        self._max_search_cells = max_search_cells

    @staticmethod
    def _evidence(
        reference: DocumentReference, *, sheet: str, cell_range: str
    ) -> EvidenceCoordinate:
        return EvidenceCoordinate(
            file_id=reference.file_id,
            sha256=reference.sha256,
            kind="SPREADSHEET",
            sheet=sheet,
            cell_range=cell_range,
        )

    def _sheets(self, reference: DocumentReference) -> list[_Sheet]:
        payload = verified_bytes(self._storage, reference, max_bytes=self._max_file_bytes)
        extension = Path(reference.filename).suffix.casefold()
        try:
            if extension == ".xlsx":
                workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=False)
                return [_OpenpyxlSheet(sheet) for sheet in workbook.worksheets]
            if extension == ".xls":
                book = xlrd.open_workbook(file_contents=payload, on_demand=True)
                return [
                    _MatrixSheet(
                        name=sheet.name,
                        rows=[
                            [sheet.cell_value(row, column) for column in range(sheet.ncols)]
                            for row in range(sheet.nrows)
                        ],
                    )
                    for sheet in book.sheets()
                ]
            if extension in {".csv", ".tsv"}:
                text = self._decode_csv(payload)
                sample = text[:8192]
                delimiter = "\t" if extension == ".tsv" else self._delimiter(sample)
                rows: list[list[object]] = [
                    list(row) for row in csv.reader(StringIO(text), delimiter=delimiter)
                ]
                return [_MatrixSheet("Sheet1", rows)]
        except DocumentToolError:
            raise
        except Exception as exc:
            raise DocumentToolError("DOCUMENT_CORRUPT", "spreadsheet could not be opened") from exc
        raise DocumentToolError("DOCUMENT_UNSUPPORTED", "spreadsheet format is unsupported")

    @staticmethod
    def _decode_csv(payload: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return payload.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise DocumentToolError("DOCUMENT_UNSUPPORTED", "CSV encoding is unsupported")

    @staticmethod
    def _delimiter(sample: str) -> str:
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            return ","

    @staticmethod
    def _sheet(sheets: list[_Sheet], name: str) -> _Sheet:
        for sheet in sheets:
            if sheet.name == name:
                return sheet
        raise DocumentToolError("DOCUMENT_RANGE_ERROR", "spreadsheet sheet does not exist")

    def list_sheets(self, reference: DocumentReference) -> SheetList:
        sheets = self._sheets(reference)
        return SheetList(
            sheets=tuple(
                SheetInfo(
                    name=sheet.name,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    evidence=self._evidence(
                        reference,
                        sheet=sheet.name,
                        cell_range=(
                            f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                            if sheet.max_row and sheet.max_column
                            else "A1:A1"
                        ),
                    ),
                )
                for sheet in sheets
            )
        )

    def read_range(
        self, reference: DocumentReference, *, sheet_name: str, cell_range: str
    ) -> SheetRangeResult:
        try:
            min_column, min_row, max_column, max_row = range_boundaries(cell_range.upper())
        except ValueError as exc:
            raise DocumentToolError("DOCUMENT_RANGE_ERROR", "invalid spreadsheet range") from exc
        count = (max_row - min_row + 1) * (max_column - min_column + 1)
        if count > self._max_cells:
            raise DocumentToolError("DOCUMENT_LIMIT", "spreadsheet range exceeds cell limit")
        sheet = self._sheet(self._sheets(reference), sheet_name)
        if max_row > max(sheet.max_row, 1) or max_column > max(sheet.max_column, 1):
            raise DocumentToolError("DOCUMENT_RANGE_ERROR", "spreadsheet range is out of bounds")
        normalized = (
            f"{get_column_letter(min_column)}{min_row}:{get_column_letter(max_column)}{max_row}"
        )
        cells = []
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                value = sheet.value(row, column)
                coordinate = f"{get_column_letter(column)}{row}"
                cells.append(
                    CellValue(
                        coordinate=coordinate,
                        value=None if value is None else str(value),
                        formula=sheet.formula(row, column),
                    )
                )
        return SheetRangeResult(
            sheet=sheet_name,
            cell_range=normalized,
            cells=tuple(cells),
            evidence=self._evidence(reference, sheet=sheet_name, cell_range=normalized),
        )

    def search_cells(
        self, reference: DocumentReference, *, query: str, max_matches: int = 50
    ) -> CellSearchResult:
        if not query.strip() or not 1 <= max_matches <= 100:
            raise DocumentToolError("DOCUMENT_LIMIT", "invalid spreadsheet search limits")
        needle = query.casefold()
        inspected = 0
        matches: list[CellSearchMatch] = []
        for sheet in self._sheets(reference):
            for row in range(1, sheet.max_row + 1):
                for column in range(1, sheet.max_column + 1):
                    inspected += 1
                    if inspected > self._max_search_cells:
                        raise DocumentToolError(
                            "DOCUMENT_LIMIT", "spreadsheet search exceeds cell limit"
                        )
                    value = sheet.value(row, column)
                    if value is None or needle not in str(value).casefold():
                        continue
                    coordinate = f"{get_column_letter(column)}{row}"
                    matches.append(
                        CellSearchMatch(
                            sheet=sheet.name,
                            coordinate=coordinate,
                            value=str(value),
                            evidence=self._evidence(
                                reference, sheet=sheet.name, cell_range=coordinate
                            ),
                        )
                    )
                    if len(matches) >= max_matches:
                        return CellSearchResult(matches=tuple(matches))
        return CellSearchResult(matches=tuple(matches))
